import streamlit as st
import pandas as pd
import re
import io
import csv
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================================
# PAGE HEADER
# ==========================================================
st.markdown(
    "<h2 style='text-align:center;'>Escondida — Ahorros Report Generator</h2>",
    unsafe_allow_html=True,
)
st.markdown(
    "<p style='text-align:center; color:gray;'>"
    "Upload QAQC files, clean, and generate the structured Excel report with formulas.</p>",
    unsafe_allow_html=True,
)
st.markdown("---")

if st.button("Back to Menu", key="back_ahorros"):
    st.session_state.page = "dashboard"
    st.rerun()

# ==========================================================
# CLEANING FUNCTIONS  (shared logic with ES_QAQC)
# ==========================================================

def read_csv_smart(file_obj):
    sample_bytes = file_obj.read(8192)
    file_obj.seek(0)
    encodings_to_try = ("utf-8", "cp1252", "latin1", "iso-8859-1")
    delimiters = [",", ";", "\t", "|"]
    for enc in encodings_to_try:
        try:
            text = sample_bytes.decode(enc, errors="replace")
        except Exception:
            continue
        sep = None
        try:
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(text, delimiters="".join(delimiters))
            sep = dialect.delimiter
        except Exception:
            if text.count(";") > text.count(","):
                sep = ";"
            elif "\t" in text:
                sep = "\t"
            elif "|" in text:
                sep = "|"
            else:
                sep = ","
        try:
            file_obj.seek(0)
            return pd.read_csv(file_obj, sep=sep, engine="python", encoding=enc)
        except Exception:
            file_obj.seek(0)
            continue
    file_obj.seek(0)
    return pd.read_csv(file_obj, sep=None, engine="python", encoding="latin1")


def _replace_dash_with_na(series):
    if series is None:
        return series
    return series.replace(["-", " -", "- ", "\u2014", "\u2013", "", "\xa0"], pd.NA)


def _to_numeric(series):
    return pd.to_numeric(_replace_dash_with_na(series), errors="coerce")


def find_water_level_column(df):
    for c in df.columns:
        key = re.sub(r"\s+", "", str(c).strip().lower())
        if ("water" in key) and ("lev" in key):
            return c
    return None


def extract_level_from_blast(text):
    if pd.isna(text):
        return None
    m = re.search(r"(\d{4})", str(text))
    return int(m.group(1)) if m else None


def extract_expansion_from_blast(text):
    if pd.isna(text):
        return None
    txt = str(text).upper()
    if "N17B" in txt:
        return 170
    if "PL1S" in txt:
        return 111
    m = re.search(r"N(\d{1,2})(?![A-Z])", txt)
    if m:
        return int(m.group(1))
    m = re.search(r"PL(\d{1,2})(?![A-Z])", txt)
    if m:
        return int(m.group(1))
    m = re.search(r"(?:S|L|E)(\d{1,2})", txt)
    if m:
        return int(m.group(1))
    return None


def parse_borehole_and_grid(raw_val):
    if pd.isna(raw_val):
        return None, ""
    s = str(raw_val).strip()
    if s == "":
        return None, ""
    s = re.sub(r"\s+", "", s)
    if "_" in s:
        left, right = s.split("_", 1)
        grid = int(left) if left.isdigit() else None
        suffix = right
    else:
        grid = None
        suffix = s
    suffix_low = suffix.lower()
    if suffix_low.startswith("aux"):
        return grid, None
    m = re.match(r"^([a-z])(\d+)$", suffix_low)
    if m:
        letter, num = m.groups()
        if letter == "b":
            return grid, int("100000" + num)
        elif letter == "c":
            return grid, int("200000" + num)
        elif letter == "d":
            return grid, int(num)
        else:
            return grid, None
    if suffix_low.isdigit():
        return grid, int(suffix_low)
    return grid, None


def fill_boreholes_by_blast(df):
    def _fill_group(group):
        counter = 10000
        new_vals = []
        for v in group["Borehole"]:
            if v == "" or pd.isna(v):
                new_vals.append(counter)
                counter += 1
            else:
                new_vals.append(v)
        group["Borehole"] = new_vals
        return group
    return df.groupby("Blast", group_keys=False).apply(_fill_group)


def cross_fill_pair(df, col_a, col_b, steps_done, label):
    if col_a not in df.columns or col_b not in df.columns:
        steps_done.append(f"  {label}: columns not found ({col_a}, {col_b}).")
        return df
    df[col_a] = _replace_dash_with_na(df[col_a])
    df[col_b] = _replace_dash_with_na(df[col_b])
    df[col_a] = df[col_a].fillna(df[col_b])
    df[col_b] = df[col_b].fillna(df[col_a])
    steps_done.append(f"  Cross-filled {label}.")
    return df


def process_file(df):
    steps_done = []

    # Clean invalid Density
    if "Density" in df.columns:
        before = len(df)
        df["Density_clean"] = pd.to_numeric(df["Density"], errors="coerce")
        df = df[df["Density_clean"].notna() & (df["Density_clean"] > 0)]
        deleted = before - len(df)
        df.drop(columns=["Density_clean"], inplace=True)
        steps_done.append(f"  Density: removed {deleted} invalid rows.")

    # Remove negative coordinates
    if "Local X (Design)" in df.columns and "Local Y (Design)" in df.columns:
        before = len(df)
        df["Local X (Design)"] = _to_numeric(df["Local X (Design)"])
        df["Local Y (Design)"] = _to_numeric(df["Local Y (Design)"])
        df = df[(df["Local X (Design)"] >= 0) & (df["Local Y (Design)"] >= 0)]
        deleted = before - len(df)
        steps_done.append(f"  Coordinates: removed {deleted} rows with negatives.")

    # Level, Expansion, Grid, Borehole
    if "Blast" in df.columns:
        df["Level"] = df["Blast"].apply(extract_level_from_blast)
        df["Expansion"] = df["Blast"].apply(extract_expansion_from_blast)
        if "Borehole" in df.columns:
            grids, bores = [], []
            for v in df["Borehole"]:
                g, b = parse_borehole_and_grid(v)
                grids.append(g)
                bores.append(b if b is not None else None if v is not None else "")
            df["Grid"] = grids
            df["Borehole"] = bores
            before_invalid = len(df)
            df = df[df["Borehole"].notna()]
            deleted_invalid = before_invalid - len(df)
            df["Borehole"] = df["Borehole"].apply(lambda x: "" if x is None else x)
            df = fill_boreholes_by_blast(df)
            steps_done.append(f"  Parsed Level/Expansion/Grid/Borehole ({deleted_invalid} invalid removed).")
            cols = list(df.columns)
            for c in ["Level", "Expansion", "Grid", "Borehole"]:
                if c in cols:
                    cols.remove(c)
            if "Blast" in cols:
                idx = cols.index("Blast")
                cols[idx + 1: idx + 1] = ["Level", "Expansion", "Grid", "Borehole"]
                df = df[cols]

    # Hole Length cross-fill
    if "Hole Length (Design)" in df.columns and "Hole Length (Actual)" in df.columns:
        before = len(df)
        df = cross_fill_pair(df, "Hole Length (Design)", "Hole Length (Actual)", steps_done, "Hole Length")
        df.dropna(subset=["Hole Length (Design)", "Hole Length (Actual)"], how="all", inplace=True)
        deleted = before - len(df)
        steps_done.append(f"  Hole Length: removed {deleted} fully-empty rows.")

    # Explosive cross-fill
    if "Explosive (kg) (Design)" in df.columns and "Explosive (kg) (Actual)" in df.columns:
        before = len(df)
        df = cross_fill_pair(df, "Explosive (kg) (Design)", "Explosive (kg) (Actual)", steps_done, "Explosive")
        df.dropna(subset=["Explosive (kg) (Design)", "Explosive (kg) (Actual)"], how="all", inplace=True)
        deleted = before - len(df)
        steps_done.append(f"  Explosive: removed {deleted} fully-empty rows.")

    # Stemming cross-fill
    if "Stemming (Design)" in df.columns and "Stemming (Actual)" in df.columns:
        before = len(df)
        df = cross_fill_pair(df, "Stemming (Design)", "Stemming (Actual)", steps_done, "Stemming")
        df.dropna(subset=["Stemming (Design)", "Stemming (Actual)"], how="all", inplace=True)
        deleted = before - len(df)
        steps_done.append(f"  Stemming: removed {deleted} fully-empty rows.")

    # Water Level: dash → 0
    water_col = find_water_level_column(df)
    if water_col:
        df[water_col] = df[water_col].astype(str).str.strip()
        df[water_col] = df[water_col].replace(["-", "\u2014", "\u2013", ""], "0")
        df[water_col] = pd.to_numeric(df[water_col], errors="coerce").fillna(0)
        steps_done.append(f"  '{water_col}': dashes → 0.")

    # Asset clean
    asset_col = next((c for c in df.columns if "Asset" in c), None)
    if asset_col:
        df[asset_col] = df[asset_col].astype(str).str.extract(r"(\d+)", expand=False)
        df[asset_col] = _replace_dash_with_na(df[asset_col])
        empty_before = int(df[asset_col].isna().sum())
        if empty_before > 0 and "Grid" in df.columns:
            grid_mode = (
                df.dropna(subset=[asset_col])
                .groupby("Grid")[asset_col]
                .agg(lambda x: x.mode().iloc[0] if len(x.mode()) > 0 else pd.NA)
            )
            filled_mask = df[asset_col].isna()
            df.loc[filled_mask, asset_col] = df.loc[filled_mask, "Grid"].map(grid_mode)
        still_empty = int(df[asset_col].isna().sum())
        if still_empty > 0:
            overall_mode = df[asset_col].mode()
            if len(overall_mode) > 0:
                df[asset_col] = df[asset_col].fillna(overall_mode.iloc[0])
        steps_done.append(f"  Asset cleaned.")

    return df, steps_done


# ==========================================================
# DENSITY FIX
# ==========================================================
def fix_density(v):
    try:
        v = float(v)
    except (ValueError, TypeError):
        return v
    if 100 <= v <= 200:
        return v / 100
    if 10 <= v <= 99:
        return v / 10
    if 2 < v <= 9:
        return v / 10
    return v


# ==========================================================
# SHEET NAME CLEANER
# ==========================================================
_MONTHS_PAT = (
    r"(?:_?)(?:Enero|Febrero|Marzo|Abril|Mayo|Junio|Julio|Agosto|"
    r"Septiembre|Octubre|Noviembre|Diciembre|"
    r"Ene|Feb|Mar|Abr|May|Jun|Jul|Ago|Sep|Oct|Nov|Dic)"
)


def clean_sheet_name(filename):
    """
    QAQC_2485_PL1_5002_04_Mayo_21.xlsx  →  2485_PL1_5002_04
    """
    name = re.sub(r"\.(xlsx|xls|csv)$", "", filename, flags=re.IGNORECASE)
    name = re.sub(r"^QAQC_", "", name, flags=re.IGNORECASE)
    name = re.sub(rf"{_MONTHS_PAT}.*$", "", name, flags=re.IGNORECASE)
    name = name.strip("_ ")
    # Sanitise for Excel sheet names
    name = re.sub(r'[\\/:*?\"\[\]<>|]', "_", name)
    return name[:31] if len(name) > 31 else (name or "Sheet")


# ==========================================================
# EXCEL REPORT CONSTANTS
# ==========================================================
HEADERS = [
    "ID malla",                                       # A  (1)
    "I Pozo",                                         # B  (2)
    "Long real BC",                                   # C  (3)
    "Pasadura real BC",                               # D  (4)
    "Taco real BC",                                   # E  (5)
    "Diámetro Protocolo BC",                          # F  (6)
    "Densidad Real calculada",                        # G  (7)  formula
    "Densidad protocolo BC",                          # H  (8)
    "Kg real BC",                                     # I  (9)
    "Burden",                                         # J  (10)
    "Spacing",                                        # K  (11)
    "Area protocolo BC",                              # L  (12) formula
    "IP protocolo BC (t/m)",                          # M  (13) formula
    "Long equivalente LB según pasaduras",            # N  (14) formula
    "Pasadura LB",                                    # O  (15)
    "Taco LB",                                        # P  (16) formula
    "Diámetro Protocolo",                             # Q  (17)
    "Densidad LB",                                    # R  (18) constant
    "Kg LB",                                          # S  (19) formula
    "B LB",                                           # T  (20)
    "S LB",                                           # U  (21)
    "Tonelaje tronado LB",                            # V  (22) formula
    "Area LB",                                        # W  (23) formula
    "IP LB (ton/m)",                                  # X  (24) formula
    "kg LB eq varicion de densidad solo produccion",  # Y  (25) formula
    "Tonelaje real tronado",                          # Z  (26) formula
]

# col_number (1-based) → source column in cleaned dataframe
DATA_MAP = {
    1:  "Grid",                     # A  ID malla
    2:  "Borehole",                 # B  I Pozo
    3:  "Hole Length (Actual)",     # C  Long real BC
    4:  "Subdrill (Design)",       # D  Pasadura real BC
    5:  "Stemming (Actual)",       # E  Taco real BC
    6:  "Diameter (Design)",       # F  Diámetro Protocolo BC
    8:  "Density",                 # H  Densidad protocolo BC
    9:  "Explosive (kg) (Actual)", # I  Kg real BC
    10: "Burden (Design)",         # J  Burden
    11: "Spacing (Design)",        # K  Spacing
    15: "Subdrill (Design)",       # O  Pasadura LB  (= Subdrill)
    17: "Diameter (Design)",       # Q  Diámetro Protocolo (= F)
    20: "Burden (Design)",         # T  B LB  (= Burden)
    21: "Spacing (Design)",        # U  S LB  (= Spacing)
}

# Summary-only columns (header in row 1, formula/value in row 2 only)
SUMMARY_HEADERS = [
    "Tonelaje tronado Estatus",                          # AA (27)
    "Suma Long real BC",                                 # AB (28)
    "Suma Kg real BC",                                   # AC (29)
    "Promedio IP protocolo BC (ton/m)",                  # AD (30)
    "Promedio Kg explosivo LB",                          # AE (31)
    "Promedio IP LB (ton/m)",                            # AF (32)
    "Suma Kg LB eq. Variación densidad sólo producción", # AG (33)
    "Promedio long",                                     # AH (34)
    "% Aumento o reducción IP",                          # AI (35)
    "N° pozos BC",                                       # AJ (36)
    "N° pozos LB",                                       # AK (37)
    "Pozos ahorrados",                                   # AL (38)
    "Kg pozos ahorrados",                                # AM (39)
    "Kg ahorro por var. Densidad",                       # AN (40)
    "Total ahorro Kg",                                   # AO (41)
    "% ahorro Kg explosivo total",                       # AP (42)
    "FC real [gr/ton]",                                  # AQ (43)
    "FC LB [gr/ton]",                                    # AR (44)
    "Ahorro según FC",                                   # AS (45)
    "X80 total Proyectado",                              # AT (46)
    "K80 total Real",                                    # AU (47)
    "X80 total LB",                                      # AV (48)
    "X80 total Proyectado Mineral",                      # AW (49)
    "K80 total Real Mineral",                            # AX (50)
    "X80 total LB Mineral",                              # AY (51)
    "X80 total Proyectado Oxido",                        # AZ (52)
    "K80 total Real Oxido",                              # BA (53)
    "X80 total LB Oxido",                                # BB (54)
    "X80 total Proyectado Lastre",                       # BC (55)
    "K80 total Real Lastre",                             # BD (56)
    "X80 total LB Lastre",                               # BE (57)
    "% de material mineral",                             # BF (58)
    "% de material oxido",                               # BG (59)
    "% de material lastre",                              # BH (60)
    "Pozos totales",                                     # BI (61)
    "Pozos medidos",                                     # BJ (62)
    "KPI Mineral",                                       # BK (63)
    "KPI Oxido",                                         # BL (64)
    "KPI Lastre",                                        # BM (65)
]

# Formulas written only in row 2 (English function names)
SUMMARY_FORMULAS = {
    "Suma Long real BC": "=SUM(C:C)",
    "Suma Kg real BC": '=SUMIFS(I:I,B:B,"<5000")',
    "Promedio IP protocolo BC (ton/m)": '=AVERAGEIF(M:M,">0")',
    "Promedio Kg explosivo LB": '=AVERAGEIFS(S:S,B:B,"<5000")',
    "Promedio IP LB (ton/m)": '=AVERAGEIF(X:X,">0")',
    "Suma Kg LB eq. Variación densidad sólo producción": "=SUM(Y:Y)",
    "Promedio long": '=AVERAGEIFS(C:C,B:B,"<5000",C:C,">0")',
    "% Aumento o reducción IP": "=(AD2-AF2)/AF2",
    "N° pozos BC": "=COUNT(B:B)",
    "N° pozos LB": "=AJ2*(1+AI2)",
    "Pozos ahorrados": "=AK2-AJ2",
    "Kg pozos ahorrados": "=AL2*AE2",
    "Kg ahorro por var. Densidad": "=AG2-AC2",
    "Total ahorro Kg": "=IF(AM2<0,AN2,AM2+AN2)",
    "% ahorro Kg explosivo total": "=IF(AM2>0,AO2/(AG2+AM2),AO2/AG2)",
    "FC real [gr/ton]": "=AC2/AA2*1000",
    "FC LB [gr/ton]": "=IF(AM2>0,(AG2+AM2)/AA2*1000,AG2/AA2*1000)",
    "Ahorro según FC": "=1-AQ2/AR2",
}

# Header color mapping: header_name → (fill_hex, font_hex)
_CLR = {
    # Group 1: Real field data — Grey
    "ID malla": ("808080", "FFFFFF"),
    "I Pozo": ("808080", "FFFFFF"),
    "Long real BC": ("808080", "FFFFFF"),
    "Pasadura real BC": ("808080", "FFFFFF"),
    "Taco real BC": ("808080", "FFFFFF"),
    "Diámetro Protocolo BC": ("808080", "FFFFFF"),
    "Densidad protocolo BC": ("808080", "FFFFFF"),
    "Kg real BC": ("808080", "FFFFFF"),
    "Burden": ("808080", "FFFFFF"),
    "Spacing": ("808080", "FFFFFF"),
    # Group 2: Calculated intermediate — Dark grey
    "Densidad Real calculada": ("4D4D4D", "FFFFFF"),
    "Area protocolo BC": ("4D4D4D", "FFFFFF"),
    "IP protocolo BC (t/m)": ("4D4D4D", "FFFFFF"),
    "Long equivalente LB según pasaduras": ("4D4D4D", "FFFFFF"),
    "Kg LB": ("4D4D4D", "FFFFFF"),
    "Tonelaje tronado LB": ("4D4D4D", "FFFFFF"),
    "Area LB": ("4D4D4D", "FFFFFF"),
    "IP LB (ton/m)": ("4D4D4D", "FFFFFF"),
    "kg LB eq varicion de densidad solo produccion": ("4D4D4D", "FFFFFF"),
    "Tonelaje real tronado": ("4D4D4D", "FFFFFF"),
    "Tonelaje tronado Estatus": ("4D4D4D", "FFFFFF"),
    "Suma Long real BC": ("4D4D4D", "FFFFFF"),
    "Suma Kg real BC": ("4D4D4D", "FFFFFF"),
    "Promedio IP protocolo BC (ton/m)": ("4D4D4D", "FFFFFF"),
    "Promedio Kg explosivo LB": ("4D4D4D", "FFFFFF"),
    "Promedio IP LB (ton/m)": ("4D4D4D", "FFFFFF"),
    "Suma Kg LB eq. Variación densidad sólo producción": ("4D4D4D", "FFFFFF"),
    "Promedio long": ("4D4D4D", "FFFFFF"),
    # Group 3: Baseline design — Blue
    "Pasadura LB": ("0070C0", "FFFFFF"),
    "Taco LB": ("0070C0", "FFFFFF"),
    "Diámetro Protocolo": ("0070C0", "FFFFFF"),
    "Densidad LB": ("0070C0", "FFFF00"),  # Yellow font
    "B LB": ("0070C0", "FFFFFF"),
    "S LB": ("0070C0", "FFFFFF"),
    # Group 4: KPIs holes/quantities — Green
    "% Aumento o reducción IP": ("00B050", "FFFFFF"),
    "N° pozos BC": ("00B050", "FFFFFF"),
    "N° pozos LB": ("00B050", "FFFFFF"),
    "Pozos ahorrados": ("00B050", "FFFFFF"),
    "Kg pozos ahorrados": ("00B050", "FFFFFF"),
    # Group 5: Density savings — Magenta
    "Kg ahorro por var. Densidad": ("D86DCD", "404040"),
    # Group 6: Total savings & FC — Purple
    "Total ahorro Kg": ("782170", "FFFFFF"),
    "% ahorro Kg explosivo total": ("782170", "FFFFFF"),
    "FC real [gr/ton]": ("782170", "FFFFFF"),
    "FC LB [gr/ton]": ("782170", "FFFFFF"),
    "Ahorro según FC": ("782170", "FFFFFF"),
    # Group 7: Fragmentation & KPIs — Teal
    "X80 total Proyectado": ("009999", "FFFFFF"),
    "K80 total Real": ("009999", "FFFFFF"),
    "X80 total LB": ("009999", "FFFFFF"),
    "X80 total Proyectado Mineral": ("009999", "FFFFFF"),
    "K80 total Real Mineral": ("009999", "FFFFFF"),
    "X80 total LB Mineral": ("009999", "FFFFFF"),
    "X80 total Proyectado Oxido": ("009999", "FFFFFF"),
    "K80 total Real Oxido": ("009999", "FFFFFF"),
    "X80 total LB Oxido": ("009999", "FFFFFF"),
    "X80 total Proyectado Lastre": ("009999", "FFFFFF"),
    "K80 total Real Lastre": ("009999", "FFFFFF"),
    "X80 total LB Lastre": ("009999", "FFFFFF"),
    "% de material mineral": ("009999", "FFFFFF"),
    "% de material oxido": ("009999", "FFFFFF"),
    "% de material lastre": ("009999", "FFFFFF"),
    "Pozos totales": ("009999", "FFFFFF"),
    "Pozos medidos": ("009999", "FFFFFF"),
    "KPI Mineral": ("009999", "FFFFFF"),
    "KPI Oxido": ("009999", "FFFFFF"),
    "KPI Lastre": ("009999", "FFFFFF"),
}

# ==========================================================
# RESUMEN AHORRO CONSTANTS
# ==========================================================
MONTH_ABBR_ES = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
}

RESUMEN_HEADERS = [
    # Group 1: Identification + calculated base — Dark grey
    ("Fecha tronadura", "4D4D4D", "FFFFFF"),
    ("Expansi\u00f3n", "4D4D4D", "FFFFFF"),
    ("Banco", "4D4D4D", "FFFFFF"),
    ("ID Tronadura", "4D4D4D", "FFFFFF"),
    ("Va con X-Energy?", "4D4D4D", "FFFFFF"),
    ("Motivo", "4D4D4D", "FFFFFF"),
    ("Datos reales insertados", "4D4D4D", "FFFFFF"),
    ("Tonelaje tronado Estatus", "4D4D4D", "FFFFFF"),
    ("Suma Long real BC", "4D4D4D", "FFFFFF"),
    ("Suma Kg real BC", "4D4D4D", "FFFFFF"),
    ("Promedio IP protocolo BC (ton/m)", "4D4D4D", "FFFFFF"),
    ("Promedio Kg explosivo LB", "4D4D4D", "FFFFFF"),
    ("Promedio IP LB (ton/m)", "4D4D4D", "FFFFFF"),
    ("Suma Kg LB eq. Variaci\u00f3n densidad s\u00f3lo producci\u00f3n", "4D4D4D", "FFFFFF"),
    ("Promedio long", "4D4D4D", "FFFFFF"),
    # Group 2: KPIs — Green
    ("% Aumento o reducci\u00f3n IP", "00B050", "FFFFFF"),
    ("N\u00b0 pozos BC", "00B050", "FFFFFF"),
    ("N\u00b0 pozos LB", "00B050", "FFFFFF"),
    ("Pozos ahorrados", "00B050", "FFFFFF"),
    ("Kg pozos ahorrados", "00B050", "FFFFFF"),
    # Group 3: Density savings — Magenta
    ("Kg ahorro por var. Densidad", "D86DCD", "404040"),
    # Group 4: Total savings — Purple
    ("Total ahorro Kg", "782170", "FFFFFF"),
    ("% ahorro Kg explosivo total", "782170", "FFFFFF"),
    ("FC real [gr/ton]", "782170", "FFFFFF"),
    ("FC LB [gr/ton]", "782170", "FFFFFF"),
    ("Ahorro seg\u00fan FC", "782170", "FFFFFF"),
    # Group 5: Economics — Dark grey
    ("Metros ahorro", "404040", "FFFFFF"),
    ("Precio Explosivo (USD/ton)", "404040", "FFFFFF"),
    ("Precio perforaci\u00f3n (USD/m)", "404040", "FFFFFF"),
    ("Precio SI (USD/un)", "404040", "FFFFFF"),
    ("Ahorro reducci\u00f3n explosivos (USD)", "404040", "FFFFFF"),
    ("Ahorro perforaci\u00f3n (USD)", "404040", "FFFFFF"),
    ("Ahorro SI (USD)", "404040", "FFFFFF"),
]


def _extract_from_name(name):
    """Extract (expansion, banco) from a cleaned sheet name."""
    n = re.sub(r"^(?:BC)[_-]?", "", name, flags=re.IGNORECASE).strip("_- ")
    parts = re.split(r"[_-]+", n)
    banco = ""
    expansion = ""
    if len(parts) >= 2:
        if re.match(r"^\d{4}$", parts[0]):
            banco = parts[0]
            expansion = parts[1]
        elif re.match(r"^\d{4}$", parts[1]):
            expansion = parts[0]
            banco = parts[1]
    return expansion, banco


def write_resumen_sheet(ws, cleaned_files):
    """Write the Resumen Ahorro sheet with cross-references to blast sheets."""
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )

    # \u2500\u2500 Headers (row 1) \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    for ci, (hdr, fill_hex, font_hex) in enumerate(RESUMEN_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = Font(name="Calibri", bold=True, color=font_hex, size=10)
        c.fill = PatternFill("solid", fgColor=fill_hex)
        c.alignment = hdr_align
        c.border = thin

    # \u2500\u2500 One row per blast \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    for i, (sheet_name, _df) in enumerate(cleaned_files):
        r = i + 2
        sn = sheet_name
        sn_esc = sn.replace("'", "''")

        expansion, banco = _extract_from_name(sn)

        # A: Fecha tronadura — empty (with border)
        ws.cell(row=r, column=1).alignment = data_align
        ws.cell(row=r, column=1).border = thin

        # B: Expansi\u00f3n
        c = ws.cell(row=r, column=2, value=expansion)
        c.alignment = data_align
        c.border = thin

        # C: Banco
        c = ws.cell(row=r, column=3,
                    value=int(banco) if banco and str(banco).isdigit() else banco)
        c.alignment = data_align
        c.border = thin

        # D: ID Tronadura
        c = ws.cell(row=r, column=4, value=sn)
        c.alignment = data_align
        c.border = thin

        # E-H: empty (with borders)
        for col_1 in (5, 6, 7, 8):
            ws.cell(row=r, column=col_1).alignment = data_align
            ws.cell(row=r, column=col_1).border = thin

        # I-Z: cross-references to blast sheet summary cells
        ref_map = {
            9:  "AB2",  # Suma Long real BC
            10: "AC2",  # Suma Kg real BC
            11: "AD2",  # Promedio IP protocolo BC
            12: "AE2",  # Promedio Kg explosivo LB
            13: "AF2",  # Promedio IP LB
            14: "AG2",  # Suma Kg LB eq
            15: "AH2",  # Promedio long
            16: "AI2",  # % Aumento o reducci\u00f3n IP
            17: "AJ2",  # N\u00b0 pozos BC
            18: "AK2",  # N\u00b0 pozos LB
            19: "AL2",  # Pozos ahorrados
            20: "AM2",  # Kg pozos ahorrados
            21: "AN2",  # Kg ahorro por var. Densidad
            22: "AO2",  # Total ahorro Kg
            23: "AP2",  # % ahorro Kg explosivo total
            24: "AQ2",  # FC real [gr/ton]
            25: "AR2",  # FC LB [gr/ton]
            26: "AS2",  # Ahorro seg\u00fan FC
        }
        for col_1, cell_ref in ref_map.items():
            c = ws.cell(row=r, column=col_1,
                        value=f"='{sn_esc}'!{cell_ref}")
            c.alignment = data_align
            c.number_format = '0.00'
            c.border = thin

        # AA (27): Metros ahorro = IF(S<0, 0, S*O)
        c = ws.cell(row=r, column=27, value=f"=IF(S{r}<0,0,S{r}*O{r})")
        c.alignment = data_align
        c.number_format = '0.00'
        c.border = thin

        # AB (28): Precio Explosivo (constant)
        c = ws.cell(row=r, column=28, value=869.094318859187)
        c.alignment = data_align
        c.number_format = '0.00'
        c.border = thin

        # AC (29): Precio perforaci\u00f3n (constant)
        c = ws.cell(row=r, column=29, value=17)
        c.alignment = data_align
        c.number_format = '0.00'
        c.border = thin

        # AD (30): Precio SI (constant)
        c = ws.cell(row=r, column=30, value=35)
        c.alignment = data_align
        c.number_format = '0.00'
        c.border = thin

        # AE (31): Ahorro reducci\u00f3n explosivos = (V/1000)*AB
        c = ws.cell(row=r, column=31, value=f"=(V{r}/1000)*AB{r}")
        c.alignment = data_align
        c.number_format = '0.00'
        c.border = thin

        # AF (32): Ahorro perforaci\u00f3n = AA*AC
        c = ws.cell(row=r, column=32, value=f"=AA{r}*AC{r}")
        c.alignment = data_align
        c.number_format = '0.00'
        c.border = thin

        # AG (33): Ahorro SI = S*AD
        c = ws.cell(row=r, column=33, value=f"=S{r}*AD{r}")
        c.alignment = data_align
        c.border = thin

    # \u2500\u2500 Column widths \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
    for ci in range(1, len(RESUMEN_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.row_dimensions[1].height = 45
    ws.freeze_panes = "A2"


# ==========================================================
# EXCEL GENERATOR
# ==========================================================
def _safe_number(val):
    """Convert a value to a proper number for openpyxl."""
    if pd.isna(val) or val == "" or val is None:
        return None
    try:
        f = float(val)
        if f == int(f):
            return int(f)
        return f
    except (ValueError, TypeError):
        return val


def write_sheet(ws, df):
    """Write one blast sheet: headers + data + formulas + summary columns."""

    all_headers = HEADERS + SUMMARY_HEADERS
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )

    # ── Headers (row 1) — all columns with per-group colors ──
    for ci, hdr in enumerate(all_headers, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        fill_hex, font_hex = _CLR.get(hdr, ("808080", "FFFFFF"))
        c.font = Font(name="Calibri", bold=True, color=font_hex, size=10)
        c.fill = PatternFill("solid", fgColor=fill_hex)
        c.alignment = hdr_align
        c.border = thin

    # ── Per-row data (starting row 2) ────────────────────────
    for ri, (_, row) in enumerate(df.iterrows()):
        r = ri + 2  # Excel row

        # Data columns (values from cleaned df)
        for col_1, src in DATA_MAP.items():
            val = _safe_number(row.get(src))
            c = ws.cell(row=r, column=col_1, value=val)
            c.alignment = data_align
            c.number_format = '0.00'
            c.border = thin

        # Densidad LB (col R=18): 0.8 for 100000XX, 0.9 for 200000XX, else 1.3
        c = ws.cell(row=r, column=18,
                    value=f'=IF(LEFT(TEXT(B{r},"0"),6)="200000",0.9,IF(LEFT(TEXT(B{r},"0"),6)="100000",0.8,1.3))')
        c.alignment = data_align
        c.number_format = '0.00'
        c.border = thin

        # Formula columns
        formulas = {
            7:  f"=(I{r}/(C{r}-E{r}))/(POWER(F{r}/25.4,2)*0.507)",
            12: f"=J{r}*K{r}",
            13: f"=IF(B{r}>5000,0,(L{r}*(C{r}-D{r})*2.5)/C{r})",
            14: f"=C{r}+(O{r}-D{r})",
            16: f'=IF(F{r}=270,5,IF(F{r}=311,5.5,""))',
            19: f"=(Q{r}/2000)^2*PI()*(N{r}-P{r})*R{r}*1000",
            22: f"=T{r}*U{r}*(N{r}-O{r})",
            23: f"=T{r}*U{r}",
            24: f"=IF(B{r}>5000,0,(W{r}*(N{r}-O{r})*2.5)/N{r})",
            25: f"=IF(B{r}>5000,0,(Q{r}/2000)^2*PI()*(N{r}-P{r})*R{r}*1000)",
            26: f"=J{r}*K{r}*(C{r}-D{r})*2.5",
        }
        for col_1, formula in formulas.items():
            c = ws.cell(row=r, column=col_1, value=formula)
            c.alignment = data_align
            c.number_format = '0.00'
            c.border = thin

    # ── Summary columns (row 2 only) ─────────────────────────
    for si, hdr in enumerate(SUMMARY_HEADERS):
        col_1 = 27 + si  # starts at column AA (27)
        formula = SUMMARY_FORMULAS.get(hdr)
        if formula:
            c = ws.cell(row=2, column=col_1, value=formula)
            c.alignment = data_align
            c.number_format = '0.00'
            c.border = thin

    # ── Column widths ────────────────────────────────────────
    data_widths = {
        1: 11, 2: 12, 3: 12, 4: 13, 5: 12, 6: 16,
        7: 18, 8: 16, 9: 12, 10: 10, 11: 10,
        12: 14, 13: 16, 14: 22, 15: 13, 16: 10,
        17: 16, 18: 12, 19: 10, 20: 8, 21: 8,
        22: 17, 23: 10, 24: 14, 25: 28, 26: 17,
    }
    for ci in range(1, len(all_headers) + 1):
        w = data_widths.get(ci, 20)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 45
    ws.freeze_panes = "A2"


def generate_excel(cleaned_files):
    """Build workbook with Resumen sheet first, then one sheet per blast."""
    wb = Workbook()
    wb.remove(wb.active)

    # Resumen Ahorro summary sheet (first sheet)
    month_abbr = MONTH_ABBR_ES[date.today().month]
    ws_resumen = wb.create_sheet(title=f"Resumen Ahorro {month_abbr}")
    write_resumen_sheet(ws_resumen, cleaned_files)

    # Individual blast sheets
    for sheet_name, df in cleaned_files:
        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, df)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ==========================================================
# MAIN UI
# ==========================================================
uploaded_files = st.file_uploader(
    "Upload QAQC files",
    type=["xlsx", "xls", "csv"],
    accept_multiple_files=True,
)

if uploaded_files:
    # ── 1) Clean every file ──────────────────────────────────
    cleaned_files = []
    all_steps = {}

    progress = st.progress(0, text="Cleaning files…")
    for i, uf in enumerate(uploaded_files):
        fname = uf.name.lower()
        df = read_csv_smart(uf) if fname.endswith(".csv") else pd.read_excel(uf)

        df_clean, steps = process_file(df)

        # Fix density scale
        if "Density" in df_clean.columns:
            df_clean["Density"] = df_clean["Density"].apply(fix_density)

        sheet_name = clean_sheet_name(uf.name)
        cleaned_files.append((sheet_name, df_clean))
        all_steps[uf.name] = steps
        progress.progress((i + 1) / len(uploaded_files), text=f"Cleaned {i + 1}/{len(uploaded_files)}")

    progress.empty()

    # ── Detect files where cleaning removed ALL rows ─────────
    empty_files = [(name, uf_name) for (name, df), uf_name
                   in zip(cleaned_files, [u.name for u in uploaded_files])
                   if len(df) == 0]
    cleaned_files = [(name, df) for name, df in cleaned_files if len(df) > 0]

    if empty_files:
        st.warning(
            f"**{len(empty_files)} file(s) had all rows removed by the filters** "
            f"(excluded from the report):"
        )
        for sheet_name, original_name in empty_files:
            st.markdown(f"  - `{original_name}` → sheet `{sheet_name}`")

    # ── Deduplicate sheet names ──────────────────────────────
    seen = {}
    for i, (name, df) in enumerate(cleaned_files):
        if name in seen:
            seen[name] += 1
            cleaned_files[i] = (f"{name}_{seen[name]}", df)
        else:
            seen[name] = 0

    st.success(
        f"Cleaned **{len(uploaded_files)}** file(s): "
        f"**{len(cleaned_files)}** with data, **{len(empty_files)}** empty (excluded)."
    )

    # ── 2) Processing steps ──────────────────────────────────
    with st.expander("Processing Steps", expanded=False):
        for fname, steps in all_steps.items():
            st.markdown(f"**{fname}**")
            for s in steps:
                st.markdown(s)
            st.markdown("---")

    # ── 3) Preview cleaned data ──────────────────────────────
    st.subheader("Cleaned Data Preview")
    preview_cols = [
        "Grid", "Borehole", "Hole Length (Actual)", "Subdrill (Design)",
        "Stemming (Actual)", "Diameter (Design)", "Density",
        "Explosive (kg) (Actual)", "Burden (Design)", "Spacing (Design)",
    ]
    for name, df in cleaned_files:
        with st.expander(f"{name}  ({len(df)} rows)"):
            cols_present = [c for c in preview_cols if c in df.columns]
            st.dataframe(df[cols_present].head(25), use_container_width=True)

    # ── 4) Quality check ─────────────────────────────────────
    st.markdown("---")
    st.subheader("Data Quality Check")

    if st.button("Run Quality Check", use_container_width=True):
        qc_cols = [
            "Grid", "Borehole", "Hole Length (Actual)", "Subdrill (Design)",
            "Stemming (Actual)", "Diameter (Design)", "Density",
            "Explosive (kg) (Actual)", "Burden (Design)", "Spacing (Design)",
        ]
        all_ok = True

        for name, df in cleaned_files:
            file_issues = []
            for col in qc_cols:
                if col not in df.columns:
                    file_issues.append(f"Missing column: **{col}**")
                    all_ok = False
                    continue

                col_issues = []
                # Empty / NaN
                empty_count = int(
                    df[col].isna().sum()
                    + (df[col].astype(str).str.strip() == "").sum()
                )
                if empty_count > 0:
                    col_issues.append(f"{empty_count} empty")
                    all_ok = False

                # Text in numeric columns
                non_empty = df[col].dropna().astype(str).str.strip()
                non_empty = non_empty[non_empty != ""]
                text_count = 0
                special_count = 0
                if len(non_empty) > 0:
                    text_mask = non_empty.apply(lambda x: bool(re.search(r"[A-Za-z]", str(x))))
                    text_count = int(text_mask.sum())

                    # Special characters
                    special_mask = non_empty.apply(
                        lambda x: bool(re.search(r"[^0-9eE.\-+\s]", str(x)))
                    )
                    special_count = int(special_mask.sum())

                if text_count > 0:
                    col_issues.append(f"{text_count} text cells")
                    all_ok = False
                if special_count > 0:
                    col_issues.append(f"{special_count} special chars")
                    all_ok = False

                if col_issues:
                    file_issues.append(f"**{col}**: " + " | ".join(col_issues))

            if file_issues:
                st.warning(f"**{name}**")
                for issue in file_issues:
                    st.markdown(f"  - {issue}")
            else:
                st.markdown(f"**{name}**: All columns OK")

        if all_ok:
            st.success("All files passed quality check — ready to generate the report.")

    # ── 5) Generate & download ───────────────────────────────
    st.markdown("---")
    st.subheader("Generate Ahorros Report")

    st.markdown(
        f"The report will contain **{len(cleaned_files)}** sheet(s): "
        + ", ".join(f"`{n}`" for n, _ in cleaned_files)
    )

    today = date.today().strftime("%Y-%m-%d")
    excel_buf = generate_excel(cleaned_files)

    st.download_button(
        "Download Ahorros Report (.xlsx)",
        excel_buf,
        file_name=f"Ahorros_Report_{today}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.markdown("<hr>", unsafe_allow_html=True)
    st.caption("Built by Maxam — Omar El Kendi")

else:
    st.info("Upload QAQC Excel or CSV files to begin.")
